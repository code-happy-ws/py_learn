import openpyxl

# 读取excel中的数据
# 第一步：打开工作簿
wb = openpyxl.load_workbook(r'E:\pycharm_len\py_learn\learn\office\file\python.xlsx')
# 第二步：选取表单
sh = wb['五一']
# 第三步：读取数据
# 参数 row:行  column：列
ce = sh.cell(row = 2,column = 1).value   # 读取第一行，第一列的数据
# print(ce)
# c2写入胡文爽，c3写入张振威,遍历
for row in range(2,10,2):
# for row in [2,4,6,8]:
    sh.cell(row=row,column=3).value='胡文爽'
    sh.cell(row=row+1,column=3).value='张振威'

wb.save(r'E:\pycharm_len\py_learn\learn\office\file\2python.xlsx')
